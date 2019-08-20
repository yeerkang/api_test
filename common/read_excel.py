"""
封装一个读取用例的excel类

# 实现读取用例数据

# 实现写入数据的功能

"""

import openpyxl


class Case:
    # 这个类用来存储用例，用例的数据设为对象的属性

    def __init__(self, attrs):
        """
        初始化用例
        :param attrs: zip类型-->[(key,value),(key1,value1)....]
        """
        for item in attrs:
            setattr(self, item[0], item[1])


class ReadExcel:
    """
    读取excel数据
    """
    def __init__(self,filename, sheetname):
        """
        这个是用例初始化读取对象的
        :param file_name:  文件名字  -->  str
        :param sheet_name: 表单名字  -->  str
        """
        self.filename = filename
        # 打开工作簿
        self.wb = openpyxl.load_workbook(filename)
        # 选择表单
        self.sheet = self.wb[sheetname]
    #
    # def __del__(self):
    #     # 特殊的方法，在对象销毁的之后执行
    #     # 关闭文件
    #     # openpyxl.load_workbook("E:\\untitled3\learning\py_1\data\\register_1.xlsx").close()
    #     self.wb.close()

    def read_data_line(self):
        """
        按行读取数据
        :return:  返回一个列表，列表中每个元素为一条用例
        """
        #按行获取数据转换成列表
        rows_data = list(self.sheet.rows)
        # 获取表单的表头信息
        titles = []
        for title in rows_data[0]:
            titles.append(title.value)
        # 定义一个空列表用来存储所有的用例数据
        datas = []
        for i in rows_data[1:]:
            # data用例临时存放用例数据
            data = []
            # 判断该单元格数据是否为字符串类型
            for j in i:
                if isinstance(j.value,str):
                    data.append(eval(j.value))
                else:
                    data.append(j.value)
            # 将表头和，该条数据内容，打包成一个字典，放入datas中
            case_data = dict(list(zip(titles,data)))
            datas.append(case_data)
        return datas

    def read_data_obj(self):
        """
        按行读取数据，表单所有数据
        每个用例存储在一个对象中
        :return: 返回一个列表，列表中每个元素为一个用例对象
        """
        #按行获取数据转换成列表
        rows_data = list(self.sheet.rows)
        # 获取表单的表头信息
        titles = []
        for title in rows_data[0]:
            titles.append(title.value)
        # 定义一个空列表用来存储所有的用例数据
        datas = []
        for i in rows_data[1:]:
            # data用例临时存放用例数据
            data = []
            for h in i:
                # 判断该单元格数据是否为字符串类型
                if isinstance(h.value,str):
                    data.append(eval(h.value))
                else:
                    data.append(h.value)
                #将该条数据放入datas中
                case_data = list(zip(titles,data))
                # 创建一个Cases类的对象，用来保存用例数据，
                case_data_obj = Case(case_data)
                for j in case_data:
                    setattr(case_data_obj, j[0], j[1])
                datas.append(case_data_obj)
        return datas

    def read_data_list(self,list_1):
        """
        list参数作为一个列表，传入的是指定读取数据的列，比如[1,2,3]
        每一行1、3、5列的数据读取出来就作为一条测试用例的数据，放在字典中
        所有的用例数据放在列表中进行返回
        """
        #获取最大行
        max_row = self.sheet.max_row
        # 定义一个空列表，用来存放所有用例数据
        datas = []
        # 定义一个空列表用来存放表头信息
        titles = []

        for row in range(1,max_row+1):
            # 判断是否是第一行
            if row == 1:
                for column in list_1:
                    title = self.sheet.cell(row,column).value
                    titles.append(title)
            else:
                #将数据保存在列表中
                #遍历表格中的数据帮保存起来
                data = []
                for column in list_1:
                    info = self.sheet.cell(row,column).value
                    data.append(info)
                    # 将该条数据和表头进行打包组合，
                    case_data = dict(zip(titles,data))
                    datas.append(case_data)
        return datas

    def read_data_list_obj(self,list_1=None):
        """
        list参数作为一个列表，传入的是指定读取数据的列，比如[1,2,3]
        每一行1、3、5列的数据读取出来就作为一条测试用例的数据，放在对象属性中
        所有的用例数据放在对象列表中进行返回
        """
        list_1 = eval(list_1)
        if list_1 == None:
            return self.read_data_obj()
        #获取最大的行
        max_row = self.sheet.max_row
        # 获取表单的表头信息
        titles = []
        #定义一个空列表存储所有的数据
        cases = []
        for row in range(1, max_row + 1):
            # 判断是否是第一行
            if row != 1:
                case_data = []  # 定义一个空列表，用来存放该行的数据
                for column in list_1:
                    data = self.sheet.cell(row, column).value
                    case_data.append(data)
                # 将该条数据和表头进行打包组合，
                case =  list(zip(titles,case_data))
                # 将一条用例存入一个对象中（每一列对应对象的一个属性）
                case_obj = Case(case)
                cases.append(case_obj)
            else:
                for column in list_1:
                    title = self.sheet.cell(row, column).value
                    titles.append(title)
        return cases



    def read_data_zd_row_and_column(self,row,column):
        #读取数据 按指定行和列来读取  row column
        ce = self.sheet.cell(row=row,column=column)
        return ce.value

    def write_data(self,row,column,msg):
        #写入数据
        self.sheet.cell(row=row,column=column,value=msg)
        self.wb.save(self.filename)



if __name__ == '__main__':
    r = ReadExcel('E:/api_test/data/api_cases.xlsx', 'login')
    data = r.read_data_list_obj()
    print(data)
